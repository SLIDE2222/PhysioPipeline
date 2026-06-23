#!/usr/bin/env python3
import csv
import json
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except Exception:
    load_workbook = None


def clean_header(value):
    return str(value or "").strip()


def normalize_row(mapping):
    return {clean_header(key): ("" if value is None else str(value)) for key, value in mapping.items()}


def read_csv_rows(file_path: Path):
    with file_path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        headers = [clean_header(name) for name in (reader.fieldnames or [])]
        rows = []
        for index, row in enumerate(reader, start=2):
            rows.append(
                {
                    "sheet": file_path.stem,
                    "rowNumber": index,
                    "cells": normalize_row(row),
                }
            )
        return headers, rows


def read_xlsx_rows(file_path: Path):
    if load_workbook is None:
        raise RuntimeError("openpyxl is not available for XLSX import.")

    workbook = load_workbook(filename=str(file_path), read_only=True, data_only=True)
    payload_rows = []
    headers = None

    for worksheet in workbook.worksheets:
        values = worksheet.iter_rows(values_only=True)
        try:
            first_row = next(values)
        except StopIteration:
            continue

        sheet_headers = [clean_header(value) for value in first_row]
        if not any(sheet_headers):
            continue

        headers = headers or sheet_headers

        for index, row_values in enumerate(values, start=2):
            mapping = {}
            for column_index, header in enumerate(sheet_headers):
                if not header:
                    continue
                cell_value = row_values[column_index] if column_index < len(row_values) else ""
                mapping[header] = "" if cell_value is None else str(cell_value)

            payload_rows.append(
                {
                    "sheet": worksheet.title,
                    "rowNumber": index,
                    "cells": normalize_row(mapping),
                }
            )

    return headers or [], payload_rows


def main():
    if len(sys.argv) < 2:
        raise SystemExit("Usage: extract-profile-rows-from-spreadsheet.py <csv-or-xlsx-path>")

    input_path = Path(sys.argv[1]).expanduser().resolve()
    if not input_path.exists():
        raise SystemExit(f"Spreadsheet not found: {input_path}")

    suffix = input_path.suffix.lower()
    if suffix == ".csv":
        headers, rows = read_csv_rows(input_path)
    elif suffix in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        headers, rows = read_xlsx_rows(input_path)
    else:
        raise SystemExit(f"Unsupported spreadsheet format: {input_path.suffix}")

    json.dump(
        {
            "filePath": str(input_path),
            "headers": headers,
            "rows": rows,
        },
        sys.stdout,
        ensure_ascii=False,
    )


if __name__ == "__main__":
    main()
